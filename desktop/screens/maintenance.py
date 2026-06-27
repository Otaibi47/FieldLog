import os
import threading
import customtkinter as ctk
from datetime import datetime, date, timedelta
from tkinter import filedialog, messagebox
from components.data_table import DataTable
from config import (
    BG, SURFACE, BORDER, TEXT_PRIMARY, TEXT_SECONDARY,
    ACCENT, ACCENT_LIGHT, SUCCESS, DANGER, FONT_FAMILY,
)
from screens.equipment import _field, _dropdown


# ─────────────────────────── screen ──────────────────────────────────────────

class MaintenanceScreen(ctk.CTkFrame):
    def __init__(self, master, api_client, **kwargs):
        super().__init__(master, fg_color=BG, **kwargs)
        self.api = api_client
        self._equipment_list: list[dict] = []
        self._all_logs: list[dict] = []       # full fetched set
        self._current_logs: list[dict] = []   # after filters
        self._build()

    # ── layout ────────────────────────────────────────────────────────────────

    def _build(self):
        # grid layout avoids pack ordering issues when toggling custom date row
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(5, weight=1)   # table row expands

        # ── Row 0: page header ─────────────────────────────────────────────
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.grid(row=0, column=0, sticky="ew", padx=24, pady=(28, 0))

        left = ctk.CTkFrame(hdr, fg_color="transparent")
        left.pack(side="left", fill="y")
        ctk.CTkLabel(
            left, text="Maintenance Log",
            font=ctk.CTkFont(family=FONT_FAMILY, size=20, weight="bold"),
            text_color=TEXT_PRIMARY, fg_color="transparent",
        ).pack(anchor="w")
        ctk.CTkLabel(
            left, text="Track all maintenance activity across your equipment",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13),
            text_color=TEXT_SECONDARY, fg_color="transparent",
        ).pack(anchor="w", pady=(2, 0))

        ctk.CTkButton(
            hdr, text="+ Log Maintenance",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13),
            fg_color=ACCENT, hover_color="#1E40AF", text_color="#FFFFFF",
            corner_radius=8, height=36,
            command=self._open_log_form,
        ).pack(side="right", anchor="center")

        # ── Row 1: divider ─────────────────────────────────────────────────
        ctk.CTkFrame(self, height=1, fg_color=BORDER, corner_radius=0).grid(
            row=1, column=0, sticky="ew", pady=(16, 0)
        )

        # ── Row 2: filter bar ──────────────────────────────────────────────
        fb = ctk.CTkFrame(self, fg_color="transparent")
        fb.grid(row=2, column=0, sticky="ew", padx=24, pady=(12, 0))

        # Export buttons (right — pack first so they don't get pushed off)
        ctk.CTkButton(
            fb, text="Export PDF", width=95,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=SURFACE, text_color=DANGER,
            border_width=1, border_color=BORDER,
            hover_color="#2D1515", corner_radius=6, height=32,
            command=self._export_pdf,
        ).pack(side="right", padx=(6, 0))

        ctk.CTkButton(
            fb, text="Export Excel", width=105,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=SURFACE, text_color=SUCCESS,
            border_width=1, border_color=BORDER,
            hover_color="#0F2B1F", corner_radius=6, height=32,
            command=self._export_excel,
        ).pack(side="right", padx=(6, 0))

        # Separator pip before export buttons
        ctk.CTkFrame(fb, width=1, fg_color=BORDER, corner_radius=0).pack(
            side="right", fill="y", padx=(8, 8)
        )

        # Technician search (right of filters, left of separator)
        self._tech_var = ctk.StringVar()
        ctk.CTkEntry(
            fb, textvariable=self._tech_var,
            placeholder_text="Search technician…",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13),
            fg_color=SURFACE, border_color=BORDER, border_width=1,
            corner_radius=6, height=32, width=190,
        ).pack(side="right", padx=(6, 0))
        self._tech_var.trace_add("write", lambda *_: self._apply_and_render())

        # Equipment filter
        self._eq_var = ctk.StringVar(value="All Equipment")
        self._eq_menu = ctk.CTkOptionMenu(
            fb, values=["All Equipment"],
            variable=self._eq_var,
            font=ctk.CTkFont(family=FONT_FAMILY, size=13),
            fg_color=SURFACE, button_color=BORDER, button_hover_color="#3D4F6B",
            text_color=TEXT_PRIMARY, dropdown_fg_color=SURFACE,
            dropdown_text_color=TEXT_PRIMARY, dropdown_hover_color=ACCENT_LIGHT,
            corner_radius=6, height=32, width=170,
            command=lambda _: self._apply_and_render(),
        )
        self._eq_menu.pack(side="left", padx=(0, 6))

        # Date range filter
        self._date_var = ctk.StringVar(value="All Time")
        ctk.CTkOptionMenu(
            fb,
            values=["All Time", "Last 7 Days", "Last 30 Days", "Last 90 Days", "Custom Range"],
            variable=self._date_var,
            font=ctk.CTkFont(family=FONT_FAMILY, size=13),
            fg_color=SURFACE, button_color=BORDER, button_hover_color="#3D4F6B",
            text_color=TEXT_PRIMARY, dropdown_fg_color=SURFACE,
            dropdown_text_color=TEXT_PRIMARY, dropdown_hover_color=ACCENT_LIGHT,
            corner_radius=6, height=32, width=148,
            command=self._on_date_range_change,
        ).pack(side="left")

        # ── Row 3: custom date row (hidden until "Custom Range" selected) ──
        self._custom_row = ctk.CTkFrame(self, fg_color="transparent")
        self._custom_row.grid(row=3, column=0, sticky="ew", padx=24, pady=(8, 0))
        self._custom_row.grid_remove()  # hidden; grid_remove preserves row slot

        ctk.CTkLabel(
            self._custom_row, text="From:",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            text_color=TEXT_SECONDARY, fg_color="transparent",
        ).pack(side="left", padx=(0, 4))
        self._from_entry = ctk.CTkEntry(
            self._custom_row, placeholder_text="YYYY-MM-DD",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13),
            fg_color=SURFACE, border_color=BORDER, border_width=1,
            corner_radius=6, height=32, width=130,
        )
        self._from_entry.pack(side="left", padx=(0, 12))

        ctk.CTkLabel(
            self._custom_row, text="To:",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            text_color=TEXT_SECONDARY, fg_color="transparent",
        ).pack(side="left", padx=(0, 4))
        self._to_entry = ctk.CTkEntry(
            self._custom_row, placeholder_text="YYYY-MM-DD",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13),
            fg_color=SURFACE, border_color=BORDER, border_width=1,
            corner_radius=6, height=32, width=130,
        )
        self._to_entry.pack(side="left", padx=(0, 12))

        ctk.CTkButton(
            self._custom_row, text="Apply",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=ACCENT, hover_color="#1E40AF", text_color="#FFFFFF",
            corner_radius=6, height=32, width=70,
            command=self._apply_and_render,
        ).pack(side="left")

        # ── Row 4: divider ─────────────────────────────────────────────────
        ctk.CTkFrame(self, height=1, fg_color=BORDER, corner_radius=0).grid(
            row=4, column=0, sticky="ew", pady=(12, 0)
        )

        # ── Row 5: table ───────────────────────────────────────────────────
        self._table = DataTable(
            self,
            columns=[
                ("Equipment",    190),
                ("Type",         100),
                ("Performed By", 150),
                ("Date",         110),
                ("Next Due",     110),
            ],
            height=480,
        )
        self._table.grid(row=5, column=0, sticky="nsew", padx=24, pady=(16, 24))

    # ── date range toggle ─────────────────────────────────────────────────────

    def _on_date_range_change(self, value: str):
        if value == "Custom Range":
            self._custom_row.grid()
        else:
            self._custom_row.grid_remove()
            self._apply_and_render()

    # ── data ──────────────────────────────────────────────────────────────────

    def refresh(self):
        threading.Thread(target=self._fetch_all, daemon=True).start()

    def _fetch_all(self):
        try:
            eq = self.api.get_equipment()
        except Exception:
            eq = []
        try:
            logs = self.api.get_maintenance_logs()   # fetch all; filter client-side
        except Exception:
            logs = []
        self.after(0, lambda: self._on_loaded(eq, logs))

    def _on_loaded(self, eq: list, logs: list):
        self._equipment_list = eq
        self._all_logs = logs
        names = ["All Equipment"] + [e["name"] for e in eq]
        self._eq_menu.configure(values=names)
        self._apply_and_render()

    # ── filtering ─────────────────────────────────────────────────────────────

    def _apply_and_render(self):
        result = self._all_logs[:]

        # Equipment filter
        eq_name = self._eq_var.get()
        if eq_name != "All Equipment":
            result = [l for l in result if l.get("equipment_name") == eq_name]

        # Date range filter
        dr = self._date_var.get()
        today = date.today()
        cutoff_from = cutoff_to = None

        if dr == "Last 7 Days":
            cutoff_from, cutoff_to = today - timedelta(days=7), today
        elif dr == "Last 30 Days":
            cutoff_from, cutoff_to = today - timedelta(days=30), today
        elif dr == "Last 90 Days":
            cutoff_from, cutoff_to = today - timedelta(days=90), today
        elif dr == "Custom Range":
            try:
                cutoff_from = date.fromisoformat(self._from_entry.get().strip())
                cutoff_to   = date.fromisoformat(self._to_entry.get().strip())
            except ValueError:
                cutoff_from = cutoff_to = None

        if cutoff_from and cutoff_to:
            filtered = []
            for l in result:
                raw = (l.get("performed_at") or "")[:10]
                try:
                    if cutoff_from <= date.fromisoformat(raw) <= cutoff_to:
                        filtered.append(l)
                except ValueError:
                    pass
            result = filtered

        # Technician search
        query = self._tech_var.get().strip().lower()
        if query:
            result = [l for l in result
                      if query in (l.get("performed_by") or "").lower()]

        self._current_logs = result
        self._render(result)

    def _render(self, logs: list):
        self._table.clear_rows()
        for i, log in enumerate(logs):
            date_str = (log.get("performed_at") or "")[:10] or "—"
            self._table.add_row([
                log.get("equipment_name") or "—",
                log.get("maintenance_type", "—").title(),
                log.get("performed_by", "—"),
                date_str,
                log.get("next_due_date", "—"),
            ], even=(i % 2 == 1))

    def _open_log_form(self):
        MaintenanceFormModal(self, self.api, self._equipment_list, on_save=self.refresh)

    # ── exports ───────────────────────────────────────────────────────────────

    def _ensure_pkg(self, pkg: str, install_name: str = None) -> bool:
        """Import pkg; auto-install via sys.executable if missing. Returns True on success."""
        import importlib, subprocess, sys
        try:
            importlib.import_module(pkg)
            return True
        except ImportError:
            pass
        try:
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", install_name or pkg, "-q"],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
            )
            importlib.import_module(pkg)
            return True
        except Exception as e:
            self.after(0, lambda err=str(e): messagebox.showerror(
                "Install failed",
                f"Could not install {pkg}.\n\nRun manually:\n  pip install {install_name or pkg}\n\n{err}",
            ))
            return False

    def _export_excel(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Export Maintenance Log",
            initialfile=f"FieldLog_Maintenance_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
        )
        if not path:
            return
        logs = self._current_logs[:]
        threading.Thread(target=self._do_excel, args=(path, logs), daemon=True).start()

    def _do_excel(self, path: str, logs: list):
        try:
            if not self._ensure_pkg("openpyxl"):
                return
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Maintenance Log"

            headers = ["Equipment", "Type", "Performed By", "Date",
                       "Next Due Date", "Description", "Parts Replaced"]
            ws.append(headers)

            accent_fill = PatternFill("solid", fgColor="1D4ED8")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = accent_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 20

            for i, log in enumerate(logs):
                ws.append([
                    log.get("equipment_name", ""),
                    (log.get("maintenance_type") or "").title(),
                    log.get("performed_by", ""),
                    (log.get("performed_at") or "")[:10],
                    log.get("next_due_date", ""),
                    log.get("description", ""),
                    log.get("parts_replaced", "") or "",
                ])
                if i % 2 == 1:
                    fill = PatternFill("solid", fgColor="F9FAFB")
                    for cell in ws[i + 2]:
                        cell.fill = fill

            for col in ws.columns:
                width = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(width + 4, 50)

            wb.save(path)
            self.after(0, lambda: os.startfile(path))

        except Exception as e:
            self.after(0, lambda err=str(e): messagebox.showerror("Export failed", err))

    def _export_pdf(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            title="Export Maintenance Log as PDF",
            initialfile=f"FieldLog_Maintenance_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
        )
        if not path:
            return
        logs       = self._current_logs[:]
        eq_filter  = self._eq_var.get()
        dr_filter  = self._date_var.get()
        tech_filter= self._tech_var.get().strip()
        threading.Thread(
            target=self._do_pdf,
            args=(path, logs, eq_filter, dr_filter, tech_filter),
            daemon=True,
        ).start()

    def _do_pdf(self, path: str, logs: list,
                eq_filter: str = "All Equipment",
                dr_filter: str = "All Time",
                tech_filter: str = ""):
        try:
            if not self._ensure_pkg("reportlab"):
                return

            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.units import mm
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle,
                Paragraph, Spacer, HRFlowable,
            )
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib.enums import TA_LEFT, TA_CENTER

            PAGE = landscape(A4)
            PW, PH = PAGE
            MARGIN  = 18 * mm
            HDR_H   = 16 * mm   # blue header bar height
            FOOT_H  = 12 * mm
            USABLE  = PW - 2 * MARGIN

            C_ACCENT      = colors.HexColor("#1D4ED8")
            C_ACCENT_DARK = colors.HexColor("#1E3A8A")
            C_ACCENT_MUTE = colors.HexColor("#BFDBFE")
            C_GRAY        = colors.HexColor("#6B7280")
            C_LIGHT_GRAY  = colors.HexColor("#F3F4F6")
            C_ALT_ROW     = colors.HexColor("#F9FAFB")
            C_BORDER      = colors.HexColor("#E5E7EB")
            C_TEXT        = colors.HexColor("#111827")

            # ── header / footer drawn on every page ────────────────────────
            def _draw_chrome(canvas, doc):
                canvas.saveState()

                # Blue header bar
                canvas.setFillColor(C_ACCENT)
                canvas.rect(0, PH - HDR_H, PW, HDR_H, fill=1, stroke=0)

                # Brand: "FieldLog"
                canvas.setFillColor(colors.white)
                canvas.setFont("Helvetica-Bold", 13)
                canvas.drawString(MARGIN, PH - HDR_H + 5.5 * mm, "FieldLog")

                # Separator pip
                canvas.setFillColor(C_ACCENT_MUTE)
                canvas.rect(MARGIN + 32 * mm, PH - HDR_H + 4 * mm, 0.4 * mm, 8 * mm,
                            fill=1, stroke=0)

                # Subtitle
                canvas.setFont("Helvetica", 10)
                canvas.drawString(MARGIN + 35 * mm, PH - HDR_H + 5.5 * mm,
                                  "Maintenance Report")

                # Date (right-aligned)
                canvas.setFont("Helvetica", 9)
                canvas.drawRightString(PW - MARGIN, PH - HDR_H + 5.5 * mm,
                                       date.today().strftime("%B %d, %Y"))

                # Footer rule
                canvas.setStrokeColor(C_BORDER)
                canvas.setLineWidth(0.5)
                canvas.line(MARGIN, FOOT_H - 2 * mm, PW - MARGIN, FOOT_H - 2 * mm)

                canvas.setFillColor(C_GRAY)
                canvas.setFont("Helvetica", 7.5)
                canvas.drawString(MARGIN, FOOT_H - 6 * mm, "FieldLog  -  Confidential")
                canvas.drawRightString(PW - MARGIN, FOOT_H - 6 * mm,
                                       f"Page {doc.page}")

                canvas.restoreState()

            doc = SimpleDocTemplate(
                path, pagesize=PAGE,
                leftMargin=MARGIN, rightMargin=MARGIN,
                topMargin=HDR_H + 10 * mm,
                bottomMargin=FOOT_H + 6 * mm,
            )

            # ── paragraph styles ───────────────────────────────────────────
            normal = ParagraphStyle("fl_normal", fontName="Helvetica",
                                    fontSize=8, leading=11, textColor=C_TEXT)
            bold   = ParagraphStyle("fl_bold",   fontName="Helvetica-Bold",
                                    fontSize=8, leading=11, textColor=C_TEXT)
            muted  = ParagraphStyle("fl_muted",  fontName="Helvetica",
                                    fontSize=8, leading=11, textColor=C_GRAY)

            elements: list = []

            # ── summary bar ────────────────────────────────────────────────
            active_filters = []
            if eq_filter   != "All Equipment": active_filters.append(f"Equipment: {eq_filter}")
            if dr_filter   != "All Time":      active_filters.append(f"Period: {dr_filter}")
            if tech_filter:                    active_filters.append(f"Technician: {tech_filter}")
            filter_str = "  |  ".join(active_filters) if active_filters else "All records"

            summary_data = [[
                Paragraph(f"<b>{len(logs)}</b>  records", bold),
                Paragraph(filter_str, muted),
                Paragraph(f"Generated {date.today().strftime('%d %b %Y')}", muted),
            ]]
            summary_tbl = Table(
                summary_data,
                colWidths=[USABLE * 0.15, USABLE * 0.60, USABLE * 0.25],
            )
            summary_tbl.setStyle(TableStyle([
                ("BACKGROUND",   (0, 0), (-1, -1), C_LIGHT_GRAY),
                ("ROWPADDING",   (0, 0), (-1, -1), 4),
                ("TOPPADDING",   (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
                ("LEFTPADDING",  (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("LINEBELOW",    (0, 0), (-1, -1), 0.5, C_BORDER),
                ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
            ]))
            elements.append(summary_tbl)
            elements.append(Spacer(1, 5 * mm))

            # ── data table ─────────────────────────────────────────────────
            col_widths = [
                USABLE * 0.18,   # Equipment
                USABLE * 0.10,   # Type
                USABLE * 0.13,   # Performed By
                USABLE * 0.09,   # Date
                USABLE * 0.09,   # Next Due
                USABLE * 0.41,   # Description
            ]

            hdr_style = ParagraphStyle("fl_hdr", fontName="Helvetica-Bold",
                                       fontSize=8, leading=10,
                                       textColor=colors.white)
            col_labels = ["Equipment", "Type", "Performed By",
                          "Date", "Next Due", "Description"]
            data = [[Paragraph(h, hdr_style) for h in col_labels]]

            for log in logs:
                desc = (log.get("description") or "")
                if len(desc) > 120:
                    desc = desc[:117] + "..."
                data.append([
                    Paragraph(log.get("equipment_name", "") or "", normal),
                    Paragraph((log.get("maintenance_type") or "").title(), normal),
                    Paragraph(log.get("performed_by", "") or "", normal),
                    Paragraph((log.get("performed_at") or "")[:10], normal),
                    Paragraph(log.get("next_due_date", "") or "", normal),
                    Paragraph(desc, normal),
                ])

            tbl = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
            tbl.setStyle(TableStyle([
                # Header row
                ("BACKGROUND",    (0, 0), (-1, 0),  C_ACCENT),
                ("TOPPADDING",    (0, 0), (-1, 0),  7),
                ("BOTTOMPADDING", (0, 0), (-1, 0),  7),
                ("LEFTPADDING",   (0, 0), (-1, 0),  8),
                # Body rows
                ("ROWBACKGROUNDS",(0, 1), (-1, -1),  [colors.white, C_ALT_ROW]),
                ("TOPPADDING",    (0, 1), (-1, -1),  5),
                ("BOTTOMPADDING", (0, 1), (-1, -1),  5),
                ("LEFTPADDING",   (0, 1), (-1, -1),  8),
                ("RIGHTPADDING",  (0, 0), (-1, -1),  6),
                # Grid
                ("LINEBELOW",     (0, 0), (-1, -1),  0.35, C_BORDER),
                ("VALIGN",        (0, 0), (-1, -1),  "TOP"),
                # Accent left rule on first column
                ("LINEAFTER",     (0, 0), (0, -1),   0.5,  C_BORDER),
            ]))
            elements.append(tbl)

            doc.build(elements,
                      onFirstPage=_draw_chrome,
                      onLaterPages=_draw_chrome)
            self.after(0, lambda: os.startfile(path))

        except Exception as e:
            self.after(0, lambda err=str(e): messagebox.showerror("Export failed", err))


# ─────────────────────────── modal ───────────────────────────────────────────

class MaintenanceFormModal(ctk.CTkToplevel):
    def __init__(self, master, api_client, equipment_list, on_save):
        super().__init__(master)
        self.api            = api_client
        self.equipment_list = equipment_list
        self.on_save        = on_save
        self.configure(fg_color=SURFACE)
        self.title("Log Maintenance")
        self.geometry("500x640")
        self.resizable(False, False)
        self.grab_set()
        self._build()

    def _build(self):
        ctk.CTkFrame(self, height=4, fg_color=ACCENT, corner_radius=0).pack(fill="x")

        title_bar = ctk.CTkFrame(self, fg_color=SURFACE, corner_radius=0)
        title_bar.pack(fill="x", padx=24, pady=(16, 0))
        ctk.CTkLabel(
            title_bar, text="Log Maintenance",
            font=ctk.CTkFont(family=FONT_FAMILY, size=16, weight="bold"),
            text_color=TEXT_PRIMARY, fg_color="transparent",
        ).pack(anchor="w")
        ctk.CTkLabel(
            title_bar, text="Record a completed maintenance event",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            text_color=TEXT_SECONDARY, fg_color="transparent",
        ).pack(anchor="w", pady=(2, 0))

        ctk.CTkFrame(self, height=1, fg_color=BORDER, corner_radius=0).pack(
            fill="x", pady=(14, 0)
        )

        body = ctk.CTkScrollableFrame(self, fg_color=SURFACE, corner_radius=0)
        body.pack(fill="both", expand=True)

        eq_names = [e["name"] for e in self.equipment_list]
        self._eq_var       = _dropdown(body, "Equipment",               eq_names or ["No equipment yet"], eq_names[0] if eq_names else "")
        self._type_var     = _dropdown(body, "Maintenance Type",        ["routine","corrective","emergency"], "routine")
        self._performed_by = _field(body, "Performed By",               "Technician name")
        self._description  = _field(body, "Description",                "What was done?")
        self._parts        = _field(body, "Parts Replaced (optional)",  "")
        self._performed_at = _field(
            body, "Performed At (ISO datetime)", "2024-06-15T09:00:00",
            datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        )
        self._next_due = _field(
            body, "Next Due Date (YYYY-MM-DD)", "2024-09-15",
            str(date.today() + timedelta(days=90)),
        )

        self._error = ctk.CTkLabel(
            body, text="",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            text_color=DANGER, fg_color="transparent",
        )
        self._error.pack(anchor="w", padx=24, pady=(4, 0))

        ctk.CTkFrame(self, height=1, fg_color=BORDER, corner_radius=0).pack(fill="x")
        foot = ctk.CTkFrame(self, fg_color=SURFACE, corner_radius=0)
        foot.pack(fill="x", padx=24, pady=16)

        ctk.CTkButton(
            foot, text="Cancel", width=88,
            fg_color=SURFACE, text_color=TEXT_PRIMARY,
            border_width=1, border_color=BORDER,
            hover_color=BG, corner_radius=6,
            command=self.destroy,
        ).pack(side="right", padx=(8, 0))
        ctk.CTkButton(
            foot, text="Save Log", width=110,
            fg_color=ACCENT, hover_color="#1E40AF",
            text_color="#FFFFFF", corner_radius=6,
            command=self._save,
        ).pack(side="right")

    def _save(self):
        eq_name = self._eq_var.get()
        eq_id   = next((e["id"] for e in self.equipment_list if e["name"] == eq_name), None)
        if not eq_id:
            self._error.configure(text="Please select a valid equipment item.")
            return

        data = {
            "equipment_id":     eq_id,
            "performed_by":     self._performed_by.get().strip(),
            "maintenance_type": self._type_var.get(),
            "description":      self._description.get().strip(),
            "parts_replaced":   self._parts.get().strip() or None,
            "performed_at":     self._performed_at.get().strip(),
            "next_due_date":    self._next_due.get().strip(),
        }

        def _run():
            try:
                self.api.create_maintenance_log(data)
                self.after(0, lambda: (self.on_save(), self.destroy()))
            except Exception as e:
                self.after(0, lambda err=str(e): self._error.configure(text=f"Error: {err}"))

        threading.Thread(target=_run, daemon=True).start()
